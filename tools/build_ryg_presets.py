"""Builds room presets from the RYG categories estimates spreadsheet.

WHY THIS IS A SCRIPT AND NOT TWENTY-FIVE HAND-WRITTEN PRESETS
------------------------------------------------------------
The refresh plan prices every room on the estate against one of twenty-seven
room types, and each of those types is a tab on a spreadsheet with a bill of
materials on it. Typing those out as Dart would be a few thousand lines nobody
could check against the sheet they came from - and the sheet is the document
that gets revised. Re-running this is how a revision reaches the app.

WHAT IT PRODUCES. One `.roompreset.json` per room type, holding the DEVICES
that type is built from, with the port ids taken straight out of the device
catalog so a device dropped from a preset and one dropped from the catalog are
the same device. It does NOT wire them: the sheet is a bill of materials and
says nothing about what plugs into what, and cabling invented here would be
cabling somebody has to check line by line against a room that does not exist
yet.

WHAT IS NOT A DEVICE. Half the sheet's rows are cost placeholders - Cabling,
Cooling, Misc, Screen, PC, Monitors, Speakers, Assistive Listening. They are
real money and not real boxes, so they are left off the drawing and priced from
the base-cost card, which is what that card is for.

Usage:  python tools/build_ryg_presets.py "path/to/RYG categories estimates.xlsx"
        python tools/build_ryg_presets.py <xlsx> --dry-run
"""
import collections
import io
import json
import os
import re
import sys

# ---------------------------------------------------------------------------
#  THE SHEETS THAT ARE ROOM TYPES
# ---------------------------------------------------------------------------
#  Everything else on the workbook is working: the equipment list the types are
#  priced from, the templates they were built out of, and the roll-ups.
NOT_A_ROOM_TYPE = {
    'Equipment', 'Template', 'Template-Subtractive', 'Counts',
    'Copy of Risk Level', 'Totals Summary (Named)', 'Category_Mapped',
    'Replacement Totals', 'Replacement Skipped Rows',
}

# ---------------------------------------------------------------------------
#  WHAT EACH SHEET ROW IS, IN THE CATALOG
# ---------------------------------------------------------------------------
#  Keyed by the sheet's own Type column. A model of None means "this is money,
#  not a box" - see the note above about placeholders.
CATALOG = {
    'Assitive Listening Kit': None,
    'Cabling': None,
    'Cooling': None,
    'Misc': None,
    'Monitors': None,
    'PC': None,
    'Screen': None,
    'Speakers': None,
    'Display Cart': None,
    'Power Strip: long cable': None,
    'Control Processor: +AVLAN': None,
    'Video Conferencing Bar: AIO': None,

    'Audio DSP: Basic': 'DMP 64 Plus C AT',
    'Audio DSP: Open Architecture': 'MRX7-D',
    'Camera: Audience / Framing': 'Cam570',
    'Camera: Presenter 10x': 'TR211',
    'Camera: Presenter 30x': 'TR335',
    'Capture Card: AV Bridge': 'AV Bridge 2x1',
    'Capture Card: Basic': 'BU113G2-BLACK',
    'Dante In': 'ADP-USBC-AU-2X2',
    'Display': 'FW-75EZ20L',
    # The catalog carries the DC-13 twice: once as the bare 'Document
    # Camera' with DC-13 as its part number, and once spelled out. The
    # spelled-out one is the unambiguous pick.
    'Doc Cam': 'Document Camera (DC-13)',
    'HDMI over Base-T: Tx or Rx': 'DTP HDMI 4K 330 Tx',
    'Microphone: Ceiling': 'MXA920',
    'Microphone: RF Kit': 'MXWAPXD2',
    'Network Switch': 'TL-SG108PE',
    'PDU: 1x Switched 1x Surge': 'AP7900B',
    'Projector + Mount': 'PT-VMZ62BU8',
    'Switcher: 82': 'DTP2 CrossPoint 82 IPCP SA',
    'Switcher: 84': 'DTP CrossPoint 84 4K IPCP Q SA',
    'Switcher: 108': 'DTP CrossPoint 108 4K IPCP Q SA LL',
    # The Equipment tab spells this with two spaces and one room-type
    # sheet with one. Both normalise to the single-space form.
    'Switcher: 42': 'DTP3 CrossPoint 42',
    'Touch Panel: 10"': 'TLP Pro 1025',
    'Touch Panel: 7"': 'TLP Pro 725',
    'USB Toggle': 'Toggle',
    'USB-C to HDMI + Cable': 'USB-C HD 101',
    'Wireless Mirroring': 'ShareLink Pro 2000',
    'Digital Signage Player': 'RoomCast',
    'Distribution Amp': 'DTP HD DA4 4K 330',
    'Network Button Panel': 'NBP 100',
    'Video Conferencing Bar: BYD': 'Bar BYOD',
    'Projector': 'PT-VMZ62BU8',
}

# ---------------------------------------------------------------------------
#  THE FOUR SHEETS THAT NAME THINGS THEIR OWN WAY
# ---------------------------------------------------------------------------
#  '1 Projector', '1 Proj 1 Cam 1 Mic', '2 Projector' and '2 Projector 1 Cam 1
#  Mic' were written before the equipment list settled, and say 'Switcher'
#  where the others say 'Switcher: 84'. Between them they price 171 of the 398
#  rooms on the estate, so they cannot simply be skipped.
#
#  The switcher is the one that cannot be read off the name alone, and it is
#  sized the way the canonical sheets size it: the 82 in a one-projector room,
#  the 84 in a two-projector one.
LEGACY = {
    'Switcher': {1: 'Switcher: 82', 2: 'Switcher: 84'},
    'Touch panel': 'Touch Panel: 7"',
    'Camera': 'Camera: Presenter 10x',
    'Room Mic': 'Microphone: Ceiling',
    'Microphone': 'Microphone: RF Kit',
    'Capture card': 'Capture Card: Basic',
    'Tplink': 'Network Switch',
    'PDU network': 'PDU: 1x Switched 1x Surge',
    'HDMI to USBc': 'USB-C to HDMI + Cable',
    'Assitive Listening': 'Assitive Listening Kit',
    'Receiver': 'HDMI over Base-T: Tx or Rx',
    'Transmitter': 'HDMI over Base-T: Tx or Rx',
    'Projector': 'Projector + Mount',
    'Doc Cam': 'Doc Cam',
    'Wireless Mirroring': 'Wireless Mirroring',
    'PC': 'PC',
    'Cabling': 'Cabling',
    'Misc': 'Misc',
    'Monitors': 'Monitors',
    'USB Toggle': 'USB Toggle',
    'Dante In': 'Dante In',
    # '!7 Disp' and '!suraudio' each carry one line the equipment list has
    # never had. Named here so they are reported as unmapped rather than
    # silently dropped.
    'ipl behind displays': None,
    'NAV RX and TX': None,
    'NAVigator': None,
    'surround system': None,
    '75" Display': 'Display',
    'Web-conferencing': 'Video Conferencing Bar: AIO',
    'Power Strip': 'Power Strip: long cable',
    'Switcher: 82': 'Switcher: 82',
}

# ---------------------------------------------------------------------------
#  WHERE A THING GOES IN THE ROOM
# ---------------------------------------------------------------------------
#  A preset that dropped twenty boxes on one spot would be a preset somebody
#  has to untangle before they can read it. These are the five places this
#  shop's rooms actually put equipment.
LOCATIONS = [
    ('LOC_1', 'Lectern', 'lectern'),
    ('LOC_2', 'Front wall', 'front'),
    ('LOC_3', 'Ceiling', 'ceiling'),
    ('LOC_4', 'Rack', 'rack'),
    ('LOC_5', 'Rear wall', 'rear'),
]

WHERE = {
    'Display': 'LOC_2',
    'Projector + Mount': 'LOC_3',
    'Projector': 'LOC_3',
    'Microphone: Ceiling': 'LOC_3',
    'Microphone: RF Kit': 'LOC_4',
    'Camera: Audience / Framing': 'LOC_5',
    'Camera: Presenter 10x': 'LOC_2',
    'Camera: Presenter 30x': 'LOC_2',
    'Doc Cam': 'LOC_1',
    'USB-C to HDMI + Cable': 'LOC_1',
    'Touch Panel: 7"': 'LOC_1',
    'Touch Panel: 10"': 'LOC_1',
    'Network Button Panel': 'LOC_2',
    'Digital Signage Player': 'LOC_2',
    'Video Conferencing Bar: BYD': 'LOC_2',
}
DEFAULT_LOCATION = 'LOC_4'   # the rack, which is where the rest of it lives


def norm(s):
    return re.sub(r'\s+', ' ', (s or '').strip())


def read_workbook(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    equipment = {}
    ws = wb['Equipment']
    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = (row + (None,) * 4)[:4]
        t = norm(str(cells[0])) if cells[0] else ''
        if not t or t == 'Constants':
            continue
        equipment[t] = {'make': norm(str(cells[1] or '')),
                        'model': norm(str(cells[2] or ''))}

    types = collections.OrderedDict()
    for name in wb.sheetnames:
        if name in NOT_A_ROOM_TYPE:
            continue
        items = []
        for row in wb[name].iter_rows(values_only=True):
            cells = ['' if c is None else norm(str(c)) for c in row]
            if len(cells) < 5 or not cells[0] or cells[0] == 'Type':
                continue
            try:
                qty = float(cells[4]) if cells[4] else 0
            except ValueError:
                qty = 0
            if qty > 0:
                items.append((cells[0], int(qty)))
        types[name] = items
    return equipment, types


def preset_name(sheet):
    """The name the picker shows.

    The '!' marks a one-off room on the master sheet and means nothing inside
    the app, so it is dropped; the shorthand is written out, because a picker
    full of '!suraudio 1 Proj 2 Cam Multimic' is a picker nobody can read.
    """
    spelled = {
        '!2X (separate) Display Conf':
            '2x Display Conference (separate systems)',
        '!2 Display Conf 2 Cam 1 Mic': '2 Display Conference, 2 Camera, 1 Mic',
        '!3 Projector 3 Cam Multimic': '3 Projector, 3 Camera, Multi-mic',
        '!7 Disp 2 Proj 2 Cam Multimic':
            '7 Display, 2 Projector, 2 Camera, Multi-mic',
        '!suraudio 1 Proj 2 Cam Multimic':
            'Surround Audio, 1 Projector, 2 Camera, Multi-mic',
        '!sound system only': 'Sound system only',
        '!CDL': 'CDL',
        '1 Proj 1 Cam 1 Mic': '1 Projector, 1 Camera, 1 Mic',
        '1 Proj 1 Display 2 Cam 2 Mic':
            '1 Projector, 1 Display, 2 Camera, 2 Mic',
        '1 Proj 2 Cam 2 Mic': '1 Projector, 2 Camera, 2 Mic',
        '1 Projector 1 Cam': '1 Projector, 1 Camera',
        '1 Projector 2 Cam Multimic': '1 Projector, 2 Camera, Multi-mic',
        '2 Display 1 Cam 1 Mic': '2 Display, 1 Camera, 1 Mic',
        '2 Display 2 Cam Multimic': '2 Display, 2 Camera, Multi-mic',
        '2 Projector 1 Cam 1 Mic': '2 Projector, 1 Camera, 1 Mic',
        '2 Projector 2 Cam Multimic': '2 Projector, 2 Camera, Multi-mic',
        '2 Proj 2 Rr Disp 2 Cam Multimic':
            '2 Projector, 2 Rear Display, 2 Camera, Multi-mic',
        '4 Projector 1 Cam 1 Mic': '4 Projector, 1 Camera, 1 Mic',
        '4 Projector 2 Cam Multimic': '4 Projector, 2 Camera, Multi-mic',
        '1 Display Conf Complex': '1 Display Conference (complex)',
        'Conf 1 Display': '1 Display Conference',
        'Conf 1 Display cart': '1 Display Conference (cart)',
        '1 Display signage': '1 Display signage',
        '1 Projector': '1 Projector',
        '2 Display': '2 Display',
        '2 Projector': '2 Projector',
        '4 Projector': '4 Projector',
    }
    return spelled.get(sheet, sheet.lstrip('!'))


def projector_count(sheet):
    m = re.match(r'^!?(\d+)\s*Proj', sheet)
    return int(m.group(1)) if m else 1


def resolve(sheet, raw_type):
    """The equipment-list row a sheet line means, or None when it is money."""
    t = norm(raw_type)
    if t in CATALOG:
        return t
    if t in LEGACY:
        mapped = LEGACY[t]
        if isinstance(mapped, dict):
            return mapped.get(projector_count(sheet), mapped[1])
        return mapped
    return '__unmapped__'


def build(xlsx, out_dir, dry_run):
    equipment, types = read_workbook(xlsx)
    catalog = {}
    doc = json.load(io.open('av_devices.json', encoding='utf-8'))
    for dev in doc['devices']:
        catalog[dev.get('model')] = dev

    unmapped = collections.Counter()
    missing = collections.Counter()
    written = []

    for sheet, items in types.items():
        nodes = []
        placeholders = collections.Counter()
        seq = 0
        # Laid out in columns by where they live, so a preset opens as a
        # readable drawing rather than twenty boxes on one spot.
        column = collections.Counter()
        for raw_type, qty in items:
            row = resolve(sheet, raw_type)
            if row == '__unmapped__':
                unmapped['%s :: %s' % (sheet, raw_type)] += 1
                continue
            if row is None or CATALOG.get(row) is None:
                placeholders[row or raw_type] += qty
                continue
            model = CATALOG[row]
            dev = catalog.get(model)
            if dev is None:
                missing['%s (%s)' % (model, row)] += 1
                continue
            where = WHERE.get(row, DEFAULT_LOCATION)
            for n in range(qty):
                seq += 1
                slot = column[where]
                column[where] += 1
                nodes.append({
                    'id': 'N%d' % seq,
                    'label': row if qty == 1 else '%s %d' % (row, n + 1),
                    'model': model,
                    'x': 120.0 + 220.0 * int(where[-1]),
                    'y': 120.0 + 150.0 * slot,
                    'fromConfig': False,
                    'rackUnits': dev.get('rackUnits', 0) or 0,
                    'location': where,
                    'ports': dev.get('ports', []),
                })

        used = {n['location'] for n in nodes}
        preset = {
            '__readme':
                'Built from the RYG categories estimates spreadsheet, sheet '
                '"%s". The equipment this room type is priced on; the cabling '
                'is drawn per room. Rebuild with tools/build_ryg_presets.py.'
                % sheet,
            'name': preset_name(sheet),
            'description': _describe(sheet, nodes, placeholders),
            'jackPrefix': '',
            # What the master sheet calls this room type. The line items on a
            # job name their type in the same words, so this is what lets one
            # find the preset it was priced from - see RoomPreset.sourceName.
            'sourceName': sheet,
            'locations': [
                {'id': i, 'name': n, 'zone': z}
                for (i, n, z) in LOCATIONS if i in used
            ],
            'nodes': nodes,
            'cables': [],
            'racks': [],
            'rackItems': [],
            'rackSlots': {},
            'screenSwitches': [],
        }
        stem = re.sub(r'[^\w\- ]+', '_', preset['name'])
        written.append((preset['name'], len(nodes), sorted(placeholders)))
        if not dry_run:
            os.makedirs(out_dir, exist_ok=True)
            target = os.path.join(out_dir, '%s.roompreset.json' % stem)
            io.open(target, 'w', encoding='utf-8', newline='').write(
                json.dumps(preset, indent=2, ensure_ascii=False) + '\n')

    print('%-46s %6s  %s' % ('PRESET', 'BOXES', 'PRICED BUT NOT DRAWN'))
    print('-' * 110)
    for name, count, extras in written:
        print('%-46s %6d  %s' % (name[:46], count, ', '.join(extras)))
    print('\n%d presets%s' % (len(written), '' if dry_run else ' written to %s' % out_dir))
    if missing:
        print('\nMODELS THE CATALOG DOES NOT HAVE:')
        for k, v in missing.most_common():
            print('  %-50s in %d sheet(s)' % (k, v))
    if unmapped:
        print('\nSHEET LINES NOTHING WAS MAPPED TO:')
        for k, v in unmapped.most_common():
            print('  %s' % k)


def _describe(sheet, nodes, placeholders):
    bits = ['%d device%s' % (len(nodes), '' if len(nodes) == 1 else 's')]
    if placeholders:
        bits.append('%d priced line%s with no box to draw'
                    % (len(placeholders), '' if len(placeholders) == 1 else 's'))
    return ('RYG room type "%s": %s. Cabling is drawn per room.'
            % (sheet, ', '.join(bits)))


if __name__ == '__main__':
    args = [a for a in sys.argv[1:] if not a.startswith('--')]
    if not args:
        print(__doc__)
        sys.exit(2)
    build(args[0], 'room_presets', '--dry-run' in sys.argv)
